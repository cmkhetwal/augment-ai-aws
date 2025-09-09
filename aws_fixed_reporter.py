#!/usr/bin/env python3
"""
Fixed AWS Usage Reporter with corrected SSM commands
Now properly collects real memory and disk usage data
"""

import boto3
import pandas as pd
import json
import os
from datetime import datetime, timedelta
from typing import Dict, List, Any
import logging
import warnings
from pathlib import Path
import numpy as np
from botocore.exceptions import ClientError
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import pickle
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

warnings.filterwarnings('ignore')

class AWSFixedReporter:
    def __init__(self):
        """Initialize Fixed AWS Usage Reporter"""
        self.setup_logging()
        self.today = datetime.now().strftime('%Y-%m-%d')
        self.data_dir = Path('usage_data')
        self.data_dir.mkdir(exist_ok=True)
        self.historical_file = self.data_dir / 'historical_data.pkl'
        
        # Configuration
        self.config = {
            'sender_email': 'no-reply@bamko.net',
            'recipient_email': 'cmkhetwal@hotmail.com',
            'ses_region': 'us-east-1',
            'ses_profile': 'unified',
            'ssm_timeout': 45  # Increased timeout
        }
        
        # Dynamically discover all AWS regions
        self.config['regions'] = self.get_all_regions()
        
        self.profiles = self.get_aws_profiles()
        
    def setup_logging(self):
        """Setup logging configuration"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('aws_fixed_reporter.log'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    def get_all_regions(self) -> List[str]:
        """Dynamically discover all AWS regions"""
        try:
            # Use a default session to get regions (no profile needed for this)
            ec2 = boto3.client('ec2', region_name='us-east-1')
            response = ec2.describe_regions()
            regions = [region['RegionName'] for region in response['Regions']]
            self.logger.info(f"Discovered {len(regions)} AWS regions: {', '.join(regions)}")
            return sorted(regions)  # Sort for consistent ordering
        except Exception as e:
            self.logger.error(f"Failed to discover regions: {e}")
            # Fallback to common regions if discovery fails
            fallback_regions = ['us-east-1', 'us-east-2', 'us-west-1', 'us-west-2', 'ap-south-1', 'ap-northeast-1']
            self.logger.warning(f"Using fallback regions: {fallback_regions}")
            return fallback_regions
        
    def get_aws_profiles(self) -> List[str]:
        """Get list of AWS profiles from credentials file"""
        credentials_file = Path.home() / '.aws' / 'credentials'
        profiles = []
        
        if credentials_file.exists():
            with open(credentials_file, 'r') as f:
                for line in f:
                    if line.strip().startswith('[') and line.strip().endswith(']'):
                        profile = line.strip()[1:-1]
                        profiles.append(profile)
        
        self.logger.info(f"Found AWS profiles: {profiles}")
        return profiles
    
    def get_cloudwatch_metric(self, cloudwatch, instance_id: str, metric_name: str, namespace: str = 'AWS/EC2') -> float:
        """Get CloudWatch metric"""
        try:
            response = cloudwatch.get_metric_statistics(
                Namespace=namespace,
                MetricName=metric_name,
                Dimensions=[
                    {'Name': 'InstanceId', 'Value': instance_id}
                ],
                StartTime=datetime.utcnow() - timedelta(hours=2),
                EndTime=datetime.utcnow(),
                Period=300,
                Statistics=['Average']
            )
            
            if response['Datapoints']:
                datapoints = sorted(response['Datapoints'], key=lambda x: x['Timestamp'], reverse=True)
                return round(datapoints[0]['Average'], 2)
        except Exception as e:
            self.logger.debug(f"CloudWatch metric {metric_name} failed for {instance_id}: {str(e)}")
        
        return 0
    
    def get_fixed_ssm_metrics(self, ssm_client, instance_id: str) -> Dict:
        """Fixed SSM metrics collection with proper command formatting"""
        metrics = {
            'memory_percent': 0,
            'disk_usage_avg': 0,
            'disk_details': [],
            'collection_method': 'none',
            'success': False,
            'efs_attached': False
        }
        
        try:
            # Fixed commands with proper line breaks and EFS detection
            commands = [
                # Memory usage - simple and reliable
                "free | grep '^Mem:' | awk '{printf \"%.2f\\n\", ($3/$2) * 100.0}'",
                # Disk usage - one filesystem per line
                "df | grep -E '^/dev/' | awk '{print $1 \":\" $5}' | head -5",
                # Average disk usage across all filesystems
                "df | grep -E '^/dev/' | awk '{sum += $5; count++} END {if(count > 0) printf \"%.2f\\n\", sum/count; else print \"0\\n\"}'",
                # Check for EFS mounts and get details
                "df -hT | grep efs | awk '{print $1\":\"$3\":\"$4\":\"$6}' | head -3 || echo 'NO_EFS'"
            ]
            
            self.logger.debug(f"Sending SSM commands to {instance_id}")
            
            response = ssm_client.send_command(
                InstanceIds=[instance_id],
                DocumentName="AWS-RunShellScript",
                Parameters={'commands': commands},
                TimeoutSeconds=self.config['ssm_timeout']
            )
            
            command_id = response['Command']['CommandId']
            
            # Wait longer for command completion
            max_attempts = 15  # 45 seconds total
            for attempt in range(max_attempts):
                time.sleep(3)
                
                try:
                    output = ssm_client.get_command_invocation(
                        CommandId=command_id,
                        InstanceId=instance_id
                    )
                    
                    status = output['Status']
                    
                    if status == 'Success':
                        stdout = output['StandardOutputContent'].strip()
                        if stdout:
                            # Parse the output properly
                            lines = [line.strip() for line in stdout.split('\n') if line.strip()]
                            
                            self.logger.debug(f"SSM output for {instance_id}: {lines}")
                            
                            # Parse memory (first line)
                            if len(lines) > 0:
                                try:
                                    memory_val = float(lines[0])
                                    if 0 <= memory_val <= 100:
                                        metrics['memory_percent'] = memory_val
                                        metrics['success'] = True
                                except (ValueError, IndexError):
                                    self.logger.debug(f"Failed to parse memory from: '{lines[0]}'")
                            
                            # Parse disk details (middle lines - now exclude last 2 lines for avg and efs)
                            disk_usages = []
                            end_index = len(lines) - 2 if len(lines) > 3 else len(lines) - 1
                            for i in range(1, end_index):
                                if ':' in lines[i]:
                                    try:
                                        filesystem, percent_str = lines[i].split(':', 1)
                                        percent = float(percent_str.replace('%', ''))
                                        if 0 <= percent <= 100:
                                            metrics['disk_details'].append({
                                                'filesystem': filesystem,
                                                'usage_percent': percent
                                            })
                                            disk_usages.append(percent)
                                    except (ValueError, IndexError):
                                        continue
                            
                            # Parse average disk usage (second to last line now)
                            if len(lines) > 2:
                                try:
                                    # Parse average disk usage (now it's second to last due to EFS check)
                                    avg_disk_line = lines[-2] if len(lines) > 3 else lines[-1]
                                    avg_disk = float(avg_disk_line)
                                    if 0 <= avg_disk <= 100:
                                        metrics['disk_usage_avg'] = avg_disk
                                except (ValueError, IndexError):
                                    # Fallback: calculate from individual disks
                                    if disk_usages:
                                        metrics['disk_usage_avg'] = sum(disk_usages) / len(disk_usages)
                            elif disk_usages:
                                metrics['disk_usage_avg'] = sum(disk_usages) / len(disk_usages)
                            
                            # Parse EFS mount status and details (last line)
                            if len(lines) > 3:
                                efs_line = lines[-1].strip()
                                if efs_line and efs_line != 'NO_EFS':
                                    metrics['efs_attached'] = True
                                    # Parse EFS mount details if present
                                    # Format: filesystem:size:used:percent
                                    if ':' in efs_line:
                                        try:
                                            parts = efs_line.split(':')
                                            if len(parts) >= 4:
                                                metrics['efs_details'] = {
                                                    'filesystem': parts[0],
                                                    'size': parts[1],
                                                    'used': parts[2],
                                                    'percent': parts[3]
                                                }
                                        except:
                                            pass
                                else:
                                    metrics['efs_attached'] = False
                            
                            metrics['collection_method'] = 'ssm_fixed'
                            self.logger.debug(f"SSM success for {instance_id}: Memory={metrics['memory_percent']:.2f}%, Disk={metrics['disk_usage_avg']:.2f}%")
                            break
                        
                    elif status == 'Failed':
                        error_content = output.get('StandardErrorContent', 'No error details')
                        self.logger.debug(f"SSM command failed for {instance_id}: {error_content}")
                        break
                        
                    elif status in ['InProgress', 'Pending']:
                        continue
                    else:
                        self.logger.debug(f"SSM unexpected status for {instance_id}: {status}")
                        break
                        
                except Exception as e:
                    self.logger.debug(f"Error checking SSM status for {instance_id}: {str(e)}")
                    continue
            
            else:
                self.logger.debug(f"SSM command timed out for {instance_id}")
                
        except Exception as e:
            self.logger.debug(f"SSM command failed for {instance_id}: {str(e)}")
        
        return metrics
    
    def get_ebs_volumes_info(self, ec2_client, instance_id: str) -> Dict:
        """Get detailed EBS volume information"""
        volume_info = {
            'count': 0,
            'total_size': 0,
            'sizes': [],
            'volume_ids': []
        }
        
        try:
            response = ec2_client.describe_instances(InstanceIds=[instance_id])
            if not response['Reservations']:
                return volume_info
            
            instance = response['Reservations'][0]['Instances'][0]
            
            for device in instance.get('BlockDeviceMappings', []):
                if 'Ebs' in device:
                    volume_id = device['Ebs']['VolumeId']
                    
                    vol_response = ec2_client.describe_volumes(VolumeIds=[volume_id])
                    if vol_response['Volumes']:
                        volume = vol_response['Volumes'][0]
                        size = volume['Size']
                        
                        volume_info['count'] += 1
                        volume_info['total_size'] += size
                        volume_info['sizes'].append(size)
                        volume_info['volume_ids'].append(volume_id)
                        
        except Exception as e:
            self.logger.debug(f"EBS volume info failed for {instance_id}: {str(e)}")
        
        return volume_info
    
    def collect_ec2_data(self, profile: str, region: str) -> List[Dict]:
        """Collect EC2 data with fixed SSM commands"""
        data = []
        
        try:
            session = boto3.Session(profile_name=profile, region_name=region)
            ec2_client = session.client('ec2')
            ssm_client = session.client('ssm')
            cloudwatch = session.client('cloudwatch')
            
            # Use paginator to get ALL instances without limits
            paginator = ec2_client.get_paginator('describe_instances')
            pages = paginator.paginate(
                Filters=[
                    {'Name': 'instance-state-name', 'Values': ['running']}
                ]
            )
            
            instances_processed = 0
            
            for page in pages:
                for reservation in page['Reservations']:
                    for instance in reservation['Instances']:
                        instance_id = instance['InstanceId']
                        instance_type = instance['InstanceType']
                        
                        # Get instance name from tags
                        name = 'N/A'
                        for tag in instance.get('Tags', []):
                            if tag['Key'] == 'Name':
                                name = tag['Value']
                                break
                    
                    # Get instance type details
                    vcpus = 2
                    memory_gb = 8
                    
                    try:
                        type_info = ec2_client.describe_instance_types(InstanceTypes=[instance_type])
                        if type_info['InstanceTypes']:
                            vcpus = type_info['InstanceTypes'][0]['VCpuInfo']['DefaultVCpus']
                            memory_gb = type_info['InstanceTypes'][0]['MemoryInfo']['SizeInMiB'] / 1024
                    except:
                        # Fallback estimates
                        memory_map = {
                            't3a.nano': 0.5, 't3a.micro': 1, 't3a.small': 2, 't3a.medium': 4, 't3a.large': 8, 't3a.xlarge': 16,
                            't3.nano': 0.5, 't3.micro': 1, 't3.small': 2, 't3.medium': 4, 't3.large': 8, 't3.xlarge': 16,
                            't2.nano': 0.5, 't2.micro': 1, 't2.small': 2, 't2.medium': 4, 't2.large': 8,
                            'm5.large': 8, 'm5.xlarge': 16, 'm5.2xlarge': 32, 'm5.4xlarge': 64,
                            'c5.large': 4, 'c5.xlarge': 8, 'c5.2xlarge': 16,
                            'r5.large': 16, 'r5.xlarge': 32, 'r5.2xlarge': 64
                        }
                        memory_gb = memory_map.get(instance_type, 8)
                        vcpus = max(1, int(memory_gb / 4))
                    
                    # Get CloudWatch metrics
                    cpu_util = self.get_cloudwatch_metric(cloudwatch, instance_id, 'CPUUtilization')
                    network_in = self.get_cloudwatch_metric(cloudwatch, instance_id, 'NetworkIn') / (1024*1024)
                    network_out = self.get_cloudwatch_metric(cloudwatch, instance_id, 'NetworkOut') / (1024*1024)
                    disk_read_ops = self.get_cloudwatch_metric(cloudwatch, instance_id, 'DiskReadOps')
                    disk_write_ops = self.get_cloudwatch_metric(cloudwatch, instance_id, 'DiskWriteOps')
                    
                    # Get EBS volume information
                    ebs_info = self.get_ebs_volumes_info(ec2_client, instance_id)
                    
                    # Get SSM metrics with fixed commands
                    ssm_metrics = self.get_fixed_ssm_metrics(ssm_client, instance_id)
                    
                    # Use actual SSM data if available, otherwise estimate
                    if ssm_metrics['success']:
                        memory_percent = ssm_metrics['memory_percent']
                        disk_usage_avg = ssm_metrics['disk_usage_avg']
                        collection_method = 'SSM+CloudWatch'
                    else:
                        # Fallback estimation (much more conservative)
                        if cpu_util > 70:
                            memory_percent = min(80, cpu_util * 1.1)
                        elif cpu_util > 30:
                            memory_percent = min(60, cpu_util * 1.3)
                        else:
                            memory_percent = 25
                        
                        disk_usage_avg = 0  # Can't estimate disk usage reliably
                        collection_method = 'CloudWatch+Estimated'
                    
                    # Prepare disk usage details
                    disk_usage_details = []
                    if ssm_metrics['disk_details']:
                        for disk in ssm_metrics['disk_details']:
                            disk_usage_details.append(f"{disk['filesystem']}:{disk['usage_percent']:.1f}%")
                    
                    # Add EFS details if present
                    if ssm_metrics.get('efs_attached') and ssm_metrics.get('efs_details'):
                        efs_info = ssm_metrics['efs_details']
                        disk_usage_details.append(f"EFS:{efs_info['percent']}")
                    
                    if disk_usage_details:
                        disk_usage_summary = '; '.join(disk_usage_details)
                    else:
                        disk_usage_summary = 'SSM Failed' if not ssm_metrics['success'] else 'No Data'
                    
                    # Check EFS from SSM metrics (much more reliable than security groups)
                    if ssm_metrics['efs_attached']:
                        efs_attached = 'Yes'
                    else:
                        efs_attached = 'No'
                    
                    record = {
                        'Date': self.today,
                        'Service': 'EC2',
                        'ID': instance_id,
                        'Name/Tag': name,
                        'Type': instance_type,
                        'vCPU': vcpus,
                        'RAM(Installed GiB)': round(memory_gb, 1),
                        'CPUUtilization(%)': cpu_util,
                        'RAMUtilization(%)': round(memory_percent, 2),
                        'DiskCount': max(1, ebs_info['count']),
                        'DiskTotal(GB)': max(8, ebs_info['total_size']),
                        'DiskSizes(GB)': ','.join(map(str, ebs_info['sizes'])) if ebs_info['sizes'] else '8',
                        'DiskUsage(%)': round(disk_usage_avg, 2),
                        'DiskUsageDetails': disk_usage_summary,
                        'EFS Attached': efs_attached,
                        'NetIn(MB)': round(network_in, 2),
                        'NetOut(MB)': round(network_out, 2),
                        'DiskReadIOPS': round(disk_read_ops, 2),
                        'DiskWriteIOPS': round(disk_write_ops, 2),
                        'Profile': profile,
                        'Region': region,
                        'State': instance['State']['Name'],
                        'LaunchTime': instance['LaunchTime'].strftime('%Y-%m-%d %H:%M:%S'),
                        'DataCollection': collection_method
                    }
                    
                    data.append(record)
                    
                    status_indicator = "✓ SSM" if ssm_metrics['success'] else "⚠ Est"
                    self.logger.info(f"{status_indicator} {instance_id} ({name[:30]}) - RAM: {memory_percent:.1f}%, Disk: {disk_usage_avg:.1f}%")
                    instances_processed += 1
                    
        except Exception as e:
            self.logger.error(f"Error collecting EC2 data for {profile}/{region}: {str(e)}")
        
        return data
    
    def collect_rds_data(self, profile: str, region: str) -> List[Dict]:
        """Collect RDS data (unchanged)"""
        data = []
        
        try:
            session = boto3.Session(profile_name=profile, region_name=region)
            rds_client = session.client('rds')
            cloudwatch = session.client('cloudwatch')
            
            response = rds_client.describe_db_instances()
            
            for db_instance in response['DBInstances']:
                db_id = db_instance['DBInstanceIdentifier']
                db_class = db_instance['DBInstanceClass']
                engine = db_instance['Engine']
                allocated_storage = db_instance['AllocatedStorage']
                
                cpu_utilization = self.get_rds_metric(cloudwatch, db_id, 'CPUUtilization')
                db_connections = self.get_rds_metric(cloudwatch, db_id, 'DatabaseConnections')
                read_iops = self.get_rds_metric(cloudwatch, db_id, 'ReadIOPS')
                write_iops = self.get_rds_metric(cloudwatch, db_id, 'WriteIOPS')
                freeable_memory_mb = self.get_rds_metric(cloudwatch, db_id, 'FreeableMemory') / (1024*1024)
                
                vcpus = self.estimate_rds_vcpu(db_class)
                memory_gb = self.estimate_rds_memory(db_class)
                
                memory_utilization = 0
                if memory_gb > 0 and freeable_memory_mb > 0:
                    memory_utilization = ((memory_gb * 1024 - freeable_memory_mb) / (memory_gb * 1024)) * 100
                
                record = {
                    'Date': self.today,
                    'Service': 'RDS',
                    'ID': db_id,
                    'Name/Tag': db_id,
                    'Type': db_class,
                    'vCPU': vcpus,
                    'RAM(Installed GiB)': memory_gb,
                    'CPUUtilization(%)': cpu_utilization,
                    'RAMUtilization(%)': round(memory_utilization, 2),
                    'DiskCount': 1,
                    'DiskTotal(GB)': allocated_storage,
                    'DiskSizes(GB)': str(allocated_storage),
                    'DiskUsage(%)': 'N/A',
                    'DiskUsageDetails': 'RDS Managed',
                    'EFS Attached': 'N/A',
                    'NetIn(MB)': 0,
                    'NetOut(MB)': 0,
                    'DiskReadIOPS': read_iops,
                    'DiskWriteIOPS': write_iops,
                    'Profile': profile,
                    'Region': region,
                    'Engine': engine,
                    'Status': db_instance['DBInstanceStatus'],
                    'Connections': db_connections,
                    'DataCollection': 'CloudWatch'
                }
                
                data.append(record)
                self.logger.info(f"✓ RDS {db_id} in {profile}/{region}")
                
        except Exception as e:
            self.logger.error(f"Error collecting RDS data for {profile}/{region}: {str(e)}")
        
        return data
    
    def get_rds_metric(self, cloudwatch, db_id: str, metric_name: str) -> float:
        """Get RDS CloudWatch metric"""
        try:
            response = cloudwatch.get_metric_statistics(
                Namespace='AWS/RDS',
                MetricName=metric_name,
                Dimensions=[{'Name': 'DBInstanceIdentifier', 'Value': db_id}],
                StartTime=datetime.utcnow() - timedelta(hours=2),
                EndTime=datetime.utcnow(),
                Period=300,
                Statistics=['Average']
            )
            
            if response['Datapoints']:
                datapoints = sorted(response['Datapoints'], key=lambda x: x['Timestamp'], reverse=True)
                return round(datapoints[0]['Average'], 2)
        except:
            pass
        
        return 0
    
    def estimate_rds_vcpu(self, db_class: str) -> int:
        """Estimate vCPU count for RDS instance class"""
        vcpu_map = {
            'db.t2.micro': 1, 'db.t2.small': 1, 'db.t2.medium': 2, 'db.t2.large': 2,
            'db.t3.micro': 2, 'db.t3.small': 2, 'db.t3.medium': 2, 'db.t3.large': 2,
            'db.t3.xlarge': 4, 'db.t3.2xlarge': 8,
            'db.t4g.micro': 2, 'db.t4g.small': 2, 'db.t4g.medium': 2, 'db.t4g.large': 2,
            'db.t4g.xlarge': 4, 'db.t4g.2xlarge': 8,
            'db.m5.large': 2, 'db.m5.xlarge': 4, 'db.m5.2xlarge': 8, 'db.m5.4xlarge': 16,
            'db.m6i.large': 2, 'db.m6i.xlarge': 4, 'db.m6i.2xlarge': 8,
            'db.r5.large': 2, 'db.r5.xlarge': 4, 'db.r5.2xlarge': 8, 'db.r5.4xlarge': 16,
            'db.r6i.large': 2, 'db.r6i.xlarge': 4, 'db.r6i.2xlarge': 8,
        }
        return vcpu_map.get(db_class, 2)
    
    def estimate_rds_memory(self, db_class: str) -> float:
        """Estimate memory in GB for RDS instance class"""
        memory_map = {
            'db.t2.micro': 1, 'db.t2.small': 2, 'db.t2.medium': 4, 'db.t2.large': 8,
            'db.t3.micro': 1, 'db.t3.small': 2, 'db.t3.medium': 4, 'db.t3.large': 8,
            'db.t3.xlarge': 16, 'db.t3.2xlarge': 32,
            'db.t4g.micro': 1, 'db.t4g.small': 2, 'db.t4g.medium': 4, 'db.t4g.large': 8,
            'db.t4g.xlarge': 16, 'db.t4g.2xlarge': 32,
            'db.m5.large': 8, 'db.m5.xlarge': 16, 'db.m5.2xlarge': 32, 'db.m5.4xlarge': 64,
            'db.m6i.large': 8, 'db.m6i.xlarge': 16, 'db.m6i.2xlarge': 32,
            'db.r5.large': 16, 'db.r5.xlarge': 32, 'db.r5.2xlarge': 64, 'db.r5.4xlarge': 128,
            'db.r6i.large': 16, 'db.r6i.xlarge': 32, 'db.r6i.2xlarge': 64,
        }
        return memory_map.get(db_class, 8)
    
    def collect_all_data(self) -> pd.DataFrame:
        """Collect data from all profiles and regions"""
        all_data = []
        
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = []
            
            for profile in self.profiles:
                for region in self.config['regions']:
                    futures.append(
                        executor.submit(self.collect_ec2_data, profile, region)
                    )
                    futures.append(
                        executor.submit(self.collect_rds_data, profile, region)
                    )
            
            for future in as_completed(futures, timeout=300):  # Increased timeout
                try:
                    data = future.result(timeout=30)
                    all_data.extend(data)
                except Exception as e:
                    self.logger.error(f"Error in data collection: {str(e)}")
        
        df = pd.DataFrame(all_data)
        
        # Remove duplicates within the current run (safety net)
        if not df.empty:
            initial_count = len(df)
            df = df.drop_duplicates(subset=['ID', 'Service', 'Date'], keep='last')
            df = df.reset_index(drop=True)
            final_count = len(df)
            
            if initial_count > final_count:
                self.logger.info(f"Removed {initial_count - final_count} duplicates from current run")
        
        return df
    
    def load_historical_data(self) -> pd.DataFrame:
        """Load historical data"""
        if self.historical_file.exists():
            try:
                with open(self.historical_file, 'rb') as f:
                    return pickle.load(f)
            except:
                return pd.DataFrame()
        return pd.DataFrame()
    
    def save_historical_data(self, df: pd.DataFrame):
        """Save historical data with deduplication"""
        historical = self.load_historical_data()
        
        # Convert to DataFrame if it's a list
        if isinstance(historical, list):
            historical = pd.DataFrame(historical)
            
        if not (historical.empty if hasattr(historical, 'empty') else len(historical) == 0):
            # Remove any existing entries for today to prevent duplicates
            today_str = self.today
            old_today_count = len(historical[historical['Date'] == today_str])
            historical = historical[historical['Date'] != today_str]
            
            if old_today_count > 0:
                self.logger.info(f"Removed {old_today_count} existing entries for today to prevent duplicates")
            
            # Now add today's data
            combined = pd.concat([historical, df], ignore_index=True)
            combined['Date'] = pd.to_datetime(combined['Date'], errors='coerce')
            cutoff_date = datetime.now() - timedelta(days=30)
            combined = combined[combined['Date'] >= cutoff_date]
            
            # Final deduplication by ID and Date (safety net)
            pre_dedup_count = len(combined)
            combined = combined.drop_duplicates(subset=['ID', 'Date'], keep='last')
            post_dedup_count = len(combined)
            
            if pre_dedup_count > post_dedup_count:
                self.logger.info(f"Final deduplication removed {pre_dedup_count - post_dedup_count} duplicate entries")
        else:
            combined = df
        
        with open(self.historical_file, 'wb') as f:
            pickle.dump(combined, f)
    
    def calculate_trends(self, current_df: pd.DataFrame, previous_df: pd.DataFrame) -> pd.DataFrame:
        """Calculate trends"""
        if previous_df.empty:
            current_df['CPU_Trend'] = 'N/A'
            current_df['RAM_Trend'] = 'N/A'
            current_df['Disk_Trend'] = 'N/A'
            current_df['CPU_Change'] = 0
            current_df['RAM_Change'] = 0
            current_df['Disk_Change'] = 0
            return current_df
        
        comparison_columns = ['ID', 'CPUUtilization(%)', 'RAMUtilization(%)']
        if 'DiskUsage(%)' in previous_df.columns:
            comparison_columns.append('DiskUsage(%)')
        
        merged = current_df.merge(
            previous_df[comparison_columns],
            on='ID',
            how='left',
            suffixes=('', '_prev')
        )
        
        merged['CPU_Change'] = merged['CPUUtilization(%)'] - merged['CPUUtilization(%)_prev'].fillna(0)
        merged['RAM_Change'] = merged['RAMUtilization(%)'] - merged['RAMUtilization(%)_prev'].fillna(0)
        
        if 'DiskUsage(%)_prev' in merged.columns:
            current_disk = pd.to_numeric(merged['DiskUsage(%)'], errors='coerce').fillna(0)
            prev_disk = pd.to_numeric(merged['DiskUsage(%)_prev'], errors='coerce').fillna(0)
            merged['Disk_Change'] = current_disk - prev_disk
            
            merged['Disk_Trend'] = merged['Disk_Change'].apply(
                lambda x: 'increased' if x > 5 else ('decreased' if x < -5 else 'stable')
            )
        else:
            merged['Disk_Change'] = 0
            merged['Disk_Trend'] = 'N/A'
        
        merged['CPU_Trend'] = merged['CPU_Change'].apply(
            lambda x: 'increased' if x > 5 else ('decreased' if x < -5 else 'stable')
        )
        merged['RAM_Trend'] = merged['RAM_Change'].apply(
            lambda x: 'increased' if x > 5 else ('decreased' if x < -5 else 'stable')
        )
        
        merged['CPU_Trend'].fillna('N/A', inplace=True)
        merged['RAM_Trend'].fillna('N/A', inplace=True)
        merged['Disk_Trend'].fillna('N/A', inplace=True)
        
        columns_to_drop = [col for col in merged.columns if col.endswith('_prev')]
        merged.drop(columns_to_drop, axis=1, errors='ignore', inplace=True)
        
        return merged
    
    def create_excel_report(self, df: pd.DataFrame, filename: str = None) -> str:
        """Create Excel report"""
        if filename is None:
            filename = f"aws_fixed_report_{self.today}.xlsx"
        
        columns_to_include = [
            'Date', 'Service', 'ID', 'Name/Tag', 'Type', 'vCPU', 'RAM(Installed GiB)',
            'CPUUtilization(%)', 'RAMUtilization(%)', 'DiskCount', 'DiskTotal(GB)',
            'DiskSizes(GB)', 'DiskUsage(%)', 'DiskUsageDetails', 'EFS Attached', 
            'NetIn(MB)', 'NetOut(MB)', 'DiskReadIOPS', 'DiskWriteIOPS', 'Profile', 'Region', 'DataCollection'
        ]
        
        if 'CPU_Trend' in df.columns:
            trend_columns = ['CPU_Trend', 'RAM_Trend', 'CPU_Change', 'RAM_Change']
            if 'Disk_Trend' in df.columns:
                trend_columns.extend(['Disk_Trend', 'Disk_Change'])
            columns_to_include.extend(trend_columns)
        
        columns_to_include = [col for col in columns_to_include if col in df.columns]
        report_df = df[columns_to_include].copy()
        
        # Sort: RDS instances first, then EC2 instances
        report_df['ServiceSort'] = report_df['Service'].map({'RDS': 0, 'EC2': 1})
        report_df = report_df.sort_values(['ServiceSort', 'Service', 'Name/Tag']).drop('ServiceSort', axis=1)
        report_df = report_df.reset_index(drop=True)
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            report_df.to_excel(writer, sheet_name='Usage Data', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Usage Data']
            
            from openpyxl.styles import PatternFill
            
            if 'CPU_Trend' in report_df.columns:
                green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                
                cpu_col = report_df.columns.get_loc('CPUUtilization(%)') + 1
                ram_col = report_df.columns.get_loc('RAMUtilization(%)') + 1
                
                if 'DiskUsage(%)' in report_df.columns:
                    disk_col = report_df.columns.get_loc('DiskUsage(%)') + 1
                else:
                    disk_col = None
                
                for row_idx, row in report_df.iterrows():
                    row_num = row_idx + 2
                    
                    if row.get('CPU_Trend') == 'increased':
                        worksheet.cell(row=row_num, column=cpu_col).fill = green_fill
                    elif row.get('CPU_Trend') == 'decreased':
                        worksheet.cell(row=row_num, column=cpu_col).fill = red_fill
                    
                    if row.get('RAM_Trend') == 'increased':
                        worksheet.cell(row=row_num, column=ram_col).fill = green_fill
                    elif row.get('RAM_Trend') == 'decreased':
                        worksheet.cell(row=row_num, column=ram_col).fill = red_fill
                    
                    if disk_col and row.get('Disk_Trend') == 'increased':
                        worksheet.cell(row=row_num, column=disk_col).fill = green_fill
                    elif disk_col and row.get('Disk_Trend') == 'decreased':
                        worksheet.cell(row=row_num, column=disk_col).fill = red_fill
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        self.logger.info(f"Excel report created: {filename}")
        return filename
    
    def collect_stopped_instances(self) -> List[Dict]:
        """Collect stopped/terminated EC2 instances for email report"""
        stopped_instances = []
        
        for profile in self.profiles:
            for region in self.config['regions']:
                try:
                    session = boto3.Session(profile_name=profile, region_name=region)
                    ec2_client = session.client('ec2')
                    
                    # Get non-running instances (stopped, terminated, stopping, terminating)
                    paginator = ec2_client.get_paginator('describe_instances')
                    pages = paginator.paginate(
                        Filters=[
                            {'Name': 'instance-state-name', 'Values': ['stopped', 'terminated', 'stopping', 'terminating']}
                        ]
                    )
                    
                    for page in pages:
                        for reservation in page['Reservations']:
                            for instance in reservation['Instances']:
                                instance_id = instance['InstanceId']
                                instance_type = instance['InstanceType']
                                state = instance['State']['Name']
                                
                                # Get instance name from tags
                                name = 'N/A'
                                for tag in instance.get('Tags', []):
                                    if tag['Key'] == 'Name':
                                        name = tag['Value']
                                        break
                                
                                stopped_instances.append({
                                    'ID': instance_id,
                                    'Name': name[:50],  # Truncate long names
                                    'Type': instance_type,
                                    'State': state.title(),
                                    'Profile': profile,
                                    'Region': region
                                })
                                
                except Exception as e:
                    self.logger.error(f"Error collecting stopped instances for {profile}/{region}: {str(e)}")
                    continue
        
        return stopped_instances
    
    def create_html_report(self, df: pd.DataFrame, stopped_instances: List[Dict] = None) -> str:
        """Create HTML email body"""
        # Count successful SSM collections
        ec2_df = df[df['Service'] == 'EC2']
        ssm_success = len(ec2_df[ec2_df['DataCollection'] == 'SSM+CloudWatch']) if not ec2_df.empty else 0
        estimated = len(ec2_df[ec2_df['DataCollection'] == 'CloudWatch+Estimated']) if not ec2_df.empty else 0
        
        html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #4CAF50; color: white; }}
                tr:nth-child(even) {{ background-color: #f2f2f2; }}
                .summary {{ background-color: #e7f3fe; padding: 15px; border-radius: 5px; margin: 20px 0; }}
                .success {{ background-color: #d4edda; padding: 10px; margin: 10px 0; border-left: 4px solid #28a745; }}
                .warning {{ background-color: #fff3cd; padding: 10px; margin: 10px 0; border-left: 4px solid #ffc107; }}
                .trend-up {{ color: green; }}
                .trend-down {{ color: red; }}
            </style>
        </head>
        <body>
            <h2>AWS Usage Report - {self.today}</h2>
            
            <div class="summary">
                <h3>Summary</h3>
                <ul>
                    <li><strong>Total EC2 Instances:</strong> {len(df[df['Service'] == 'EC2'])} 
                        <small>(✅ {len(df[(df['Service'] == 'EC2') & (df['DataCollection'] == 'SSM+CloudWatch')])} with SSM, 
                        ⚠️ {len(df[(df['Service'] == 'EC2') & (df['DataCollection'] == 'CloudWatch+Estimated')])} SSM failed)</small></li>
                    <li><strong>Total RDS Instances:</strong> {len(df[df['Service'] == 'RDS'])}</li>
                    <li><strong>Total Resources:</strong> {len(df)} instances analyzed</li>
                    <li><strong>Profiles Scanned:</strong> {', '.join(df['Profile'].unique())}</li>
                    <li><strong>Regions Scanned:</strong> {', '.join(df['Region'].unique())}</li>
                </ul>
            </div>
            
            <div class="{'success' if ssm_success > 0 else 'warning'}">
                <h3>🔧 Data Collection Status - FIXED!</h3>
                <ul>
                    <li>✅ Real SSM Data: <strong>{ssm_success}</strong> instances</li>
                    <li>⚠️  Estimated Data: <strong>{estimated}</strong> instances</li>
                    <li>🎯 SSM Success Rate: <strong>{(ssm_success/(ssm_success+estimated)*100) if (ssm_success+estimated) > 0 else 0:.1f}%</strong></li>
                </ul>
            </div>
        """
        
        # Add trend summary if available
        if 'CPU_Trend' in df.columns and df['CPU_Trend'].iloc[0] != 'N/A':
            cpu_increased = len(df[df['CPU_Trend'] == 'increased'])
            cpu_decreased = len(df[df['CPU_Trend'] == 'decreased'])
            ram_increased = len(df[df['RAM_Trend'] == 'increased'])
            ram_decreased = len(df[df['RAM_Trend'] == 'decreased'])
            
            html += f"""
            <div class="summary">
                <h3>📊 Daily Trends</h3>
                <ul>
                    <li>CPU: <span class="trend-up">↑ {cpu_increased}</span> | <span class="trend-down">↓ {cpu_decreased}</span></li>
                    <li>RAM: <span class="trend-up">↑ {ram_increased}</span> | <span class="trend-down">↓ {ram_decreased}</span></li>
                </ul>
            </div>
            """
        
        # Add SSM failure analysis
        ec2_df = df[df['Service'] == 'EC2']
        if not ec2_df.empty:
            ssm_failed = ec2_df[ec2_df['DataCollection'] == 'CloudWatch+Estimated']
            if len(ssm_failed) > 0:
                html += f"""
                <div class="warning">
                    <h3>⚠️ SSM Connection Issues</h3>
                    <p>The following <strong>{len(ssm_failed)} EC2 instances</strong> could not be reached via SSM for real-time metrics:</p>
                    <table style="margin: 10px 0;">
                        <tr>
                            <th style="padding: 8px; background-color: #4CAF50; border: 1px solid #ddd; color: white;">Instance ID</th>
                            <th style="padding: 8px; background-color: #4CAF50; border: 1px solid #ddd; color: white;">Name</th>
                            <th style="padding: 8px; background-color: #4CAF50; border: 1px solid #ddd; color: white;">Type</th>
                            <th style="padding: 8px; background-color: #4CAF50; border: 1px solid #ddd; color: white;">Region</th>
                            <th style="padding: 8px; background-color: #4CAF50; border: 1px solid #ddd; color: white;">Possible Cause</th>
                        </tr>
                """
                
                for _, row in ssm_failed.iterrows():
                    # Determine likely cause
                    if 'windows' in row['Name/Tag'].lower():
                        cause = "Windows instance - SSM commands different"
                    elif row['State'] != 'running':
                        cause = f"Instance state: {row['State']}"
                    else:
                        cause = "SSM agent not installed/configured"
                    
                    html += f"""
                        <tr>
                            <td>{row['ID']}</td>
                            <td>{row['Name/Tag']}</td>
                            <td>{row['Type']}</td>
                            <td>{row['Region']}</td>
                            <td>{cause}</td>
                        </tr>
                    """
                
                html += """
                    </table>
                    <p><strong>📋 Recommendations:</strong></p>
                    <ul>
                        <li>Install/update SSM agent on failed instances</li>
                        <li>Verify IAM roles have SSM permissions</li>
                        <li>Check security group allows SSM endpoints</li>
                        <li>Windows instances may need PowerShell commands</li>
                    </ul>
                </div>
                """

        # Add stopped instances section
        if stopped_instances and len(stopped_instances) > 0:
            html += f"""
            <div class="warning">
                <h3>⏸️ Non-Running EC2 Instances</h3>
                <p>Found <strong>{len(stopped_instances)} non-running instances</strong> that may be candidates for cleanup or cost optimization:</p>
                <table style="margin: 10px 0;">
                    <tr>
                        <th style="padding: 8px; background-color: #4CAF50; border: 1px solid #ddd; color: white;">Instance ID</th>
                        <th style="padding: 8px; background-color: #4CAF50; border: 1px solid #ddd; color: white;">Name</th>
                        <th style="padding: 8px; background-color: #4CAF50; border: 1px solid #ddd; color: white;">Type</th>
                        <th style="padding: 8px; background-color: #4CAF50; border: 1px solid #ddd; color: white;">State</th>
                        <th style="padding: 8px; background-color: #4CAF50; border: 1px solid #ddd; color: white;">Profile</th>
                        <th style="padding: 8px; background-color: #4CAF50; border: 1px solid #ddd; color: white;">Region</th>
                    </tr>
            """
            
            # Limit to first 20 stopped instances to avoid email clutter
            for instance in stopped_instances[:20]:
                state_color = {
                    'Stopped': '#ff9800',
                    'Terminated': '#f44336', 
                    'Stopping': '#ff5722',
                    'Terminating': '#d32f2f'
                }.get(instance['State'], '#757575')
                
                html += f"""
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd; color: #333;">{instance['ID']}</td>
                        <td style="padding: 8px; border: 1px solid #ddd; color: #333;">{instance['Name']}</td>
                        <td style="padding: 8px; border: 1px solid #ddd; color: #333;">{instance['Type']}</td>
                        <td style="padding: 8px; border: 1px solid #ddd; color: {state_color}; font-weight: bold;">{instance['State']}</td>
                        <td style="padding: 8px; border: 1px solid #ddd; color: #333;">{instance['Profile']}</td>
                        <td style="padding: 8px; border: 1px solid #ddd; color: #333;">{instance['Region']}</td>
                    </tr>
                """
            
            if len(stopped_instances) > 20:
                html += f"""
                    <tr>
                        <td colspan="6" style="padding: 8px; border: 1px solid #ddd; font-style: italic; text-align: center; color: #666;">
                            ... and {len(stopped_instances) - 20} more instances
                        </td>
                    </tr>
                """
                
            html += """
                </table>
                <p><strong>💡 Cost Optimization Tips:</strong></p>
                <ul>
                    <li><strong>Stopped instances</strong> still incur EBS storage costs</li>
                    <li><strong>Terminated instances</strong> will disappear from this list after a few hours</li>
                    <li>Consider creating AMIs before terminating if you need to preserve configurations</li>
                    <li>Review and terminate instances that are no longer needed</li>
                </ul>
            </div>
            """

        html += """
        <h3>🏆 Top 5 Resource Consumers (by CPU)</h3>
        <table>
            <tr>
                <th>Resource ID</th>
                <th>Name</th>
                <th>Type</th>
                <th>CPU %</th>
                <th>RAM %</th>
                <th>Disk %</th>
                <th>Data Source</th>
            </tr>
        """
        
        top_cpu = df.nlargest(min(5, len(df)), 'CPUUtilization(%)')
        for _, row in top_cpu.iterrows():
            disk_usage = row.get('DiskUsage(%)', 'N/A')
            if isinstance(disk_usage, (int, float)):
                disk_display = f"{disk_usage:.1f}%"
            else:
                disk_display = str(disk_usage)
            
            data_source = row.get('DataCollection', 'Unknown')
            if 'SSM' in data_source:
                source_display = '✅ SSM+CW'
                source_class = 'trend-up'
            else:
                source_display = '⚠️ Est+CW'
                source_class = 'trend-down'
            
            html += f"""
            <tr>
                <td>{row['ID']}</td>
                <td>{row['Name/Tag']}</td>
                <td>{row['Type']}</td>
                <td>{row['CPUUtilization(%)']:.1f}%</td>
                <td>{row['RAMUtilization(%)']:.1f}%</td>
                <td>{disk_display}</td>
                <td class="{source_class}">{source_display}</td>
            </tr>
            """
        
        html += """
            </table>
            <p><em>🎯 Fixed report with real SSM data collection attached as Excel file!</em></p>
        </body>
        </html>
        """
        
        return html
    
    def analyze_monthly_trends(self) -> Dict:
        """Analyze monthly trends and generate scaling recommendations"""
        historical = self.load_historical_data()
        
        if historical.empty if hasattr(historical, 'empty') else len(historical) == 0:
            self.logger.warning("No historical data available for monthly analysis")
            return {'recommendations': [], 'analysis': {}, 'summary': {}}
        
        # Convert to DataFrame if it's a list
        if isinstance(historical, list):
            historical = pd.DataFrame(historical)
        
        # Ensure Date column is datetime
        historical['Date'] = pd.to_datetime(historical['Date'], errors='coerce')
        
        # Get last 30 days of data
        cutoff_date = datetime.now() - timedelta(days=30)
        recent_data = historical[historical['Date'] >= cutoff_date].copy()
        
        if recent_data.empty:
            self.logger.warning("Insufficient recent data for monthly analysis")
            return {'recommendations': [], 'analysis': {}, 'summary': {}}
        
        recommendations = []
        analysis_summary = {
            'total_instances': len(recent_data['ID'].unique()),
            'analysis_period_days': len(recent_data['Date'].unique()),
            'data_points': len(recent_data)
        }
        
        # Analyze each unique instance
        for instance_id in recent_data['ID'].unique():
            instance_data = recent_data[recent_data['ID'] == instance_id].copy()
            
            if len(instance_data) < 7:  # Need at least a week of data
                continue
            
            # Get basic instance info
            latest_record = instance_data.sort_values('Date').iloc[-1]
            service = latest_record['Service']
            instance_type = latest_record['Type']
            name = latest_record['Name/Tag']
            
            # Calculate statistics
            cpu_stats = self._calculate_utilization_stats(instance_data, 'CPUUtilization(%)')
            ram_stats = self._calculate_utilization_stats(instance_data, 'RAMUtilization(%)')
            
            # Get disk stats if available
            disk_stats = None
            if 'DiskUsage(%)' in instance_data.columns:
                disk_data = pd.to_numeric(instance_data['DiskUsage(%)'], errors='coerce')
                if not disk_data.isna().all():
                    disk_stats = self._calculate_utilization_stats(instance_data, 'DiskUsage(%)')
            
            # Generate recommendation
            recommendation = self._generate_instance_recommendation(
                instance_id, name, service, instance_type, 
                cpu_stats, ram_stats, disk_stats, len(instance_data)
            )
            
            if recommendation:
                recommendations.append(recommendation)
        
        # Generate overall analysis
        monthly_analysis = self._generate_monthly_summary(recent_data, recommendations)
        
        return {
            'recommendations': recommendations,
            'analysis': monthly_analysis,
            'summary': analysis_summary
        }
    
    def _calculate_utilization_stats(self, data: pd.DataFrame, column: str) -> Dict:
        """Calculate utilization statistics for a metric"""
        values = pd.to_numeric(data[column], errors='coerce').dropna()
        
        if values.empty:
            return {'avg': 0, 'max': 0, 'min': 0, 'p95': 0, 'p90': 0, 'low_usage_days': 0}
        
        stats = {
            'avg': round(values.mean(), 2),
            'max': round(values.max(), 2),
            'min': round(values.min(), 2),
            'p95': round(values.quantile(0.95), 2),
            'p90': round(values.quantile(0.90), 2),
            'low_usage_days': len(values[values < 20]),
            'high_usage_days': len(values[values > 80]),
            'total_days': len(values)
        }
        
        stats['low_usage_percentage'] = round((stats['low_usage_days'] / stats['total_days']) * 100, 1)
        stats['high_usage_percentage'] = round((stats['high_usage_days'] / stats['total_days']) * 100, 1)
        
        return stats
    
    def _generate_instance_recommendation(self, instance_id: str, name: str, service: str, 
                                        instance_type: str, cpu_stats: Dict, ram_stats: Dict, 
                                        disk_stats: Dict, data_points: int) -> Dict:
        """Generate scaling recommendation for a single instance"""
        recommendation = {
            'instance_id': instance_id,
            'name': name,
            'service': service,
            'current_type': instance_type,
            'data_points': data_points,
            'cpu_stats': cpu_stats,
            'ram_stats': ram_stats,
            'disk_stats': disk_stats,
            'action': 'monitor',
            'reason': 'Usage within normal range',
            'confidence': 'medium',
            'priority': 'low',
            'estimated_monthly_savings': 0,
            'estimated_monthly_cost_increase': 0,
            'recommended_type': instance_type
        }
        
        # Downsize recommendations (high confidence scenarios)
        if (cpu_stats['avg'] < 15 and cpu_stats['p95'] < 30 and 
            ram_stats['avg'] < 25 and ram_stats['p95'] < 50 and
            cpu_stats['low_usage_percentage'] > 70):
            
            recommended_type = self._get_smaller_instance_type(instance_type, service)
            if recommended_type != instance_type:
                savings = self._estimate_cost_savings(instance_type, recommended_type, service)
                recommendation.update({
                    'action': 'downsize',
                    'recommended_type': recommended_type,
                    'reason': f'Consistently low utilization: CPU avg {cpu_stats["avg"]}%, RAM avg {ram_stats["avg"]}%',
                    'confidence': 'high',
                    'priority': 'high' if savings > 100 else 'medium',
                    'estimated_monthly_savings': savings
                })
        
        # Upsize recommendations (medium to high confidence)
        elif (cpu_stats['avg'] > 75 and cpu_stats['p95'] > 90 and 
              cpu_stats['high_usage_percentage'] > 30):
            
            recommended_type = self._get_larger_instance_type(instance_type, service)
            if recommended_type != instance_type:
                cost_increase = self._estimate_cost_increase(instance_type, recommended_type, service)
                recommendation.update({
                    'action': 'upsize',
                    'recommended_type': recommended_type,
                    'reason': f'High CPU utilization: avg {cpu_stats["avg"]}%, 95th percentile {cpu_stats["p95"]}%',
                    'confidence': 'high',
                    'priority': 'high',
                    'estimated_monthly_cost_increase': cost_increase
                })
        
        # RAM-based recommendations
        elif (ram_stats['avg'] > 85 and ram_stats['p95'] > 95 and 
              ram_stats['high_usage_percentage'] > 25):
            
            recommended_type = self._get_memory_optimized_type(instance_type, service)
            if recommended_type != instance_type:
                cost_increase = self._estimate_cost_increase(instance_type, recommended_type, service)
                recommendation.update({
                    'action': 'optimize_memory',
                    'recommended_type': recommended_type,
                    'reason': f'High RAM utilization: avg {ram_stats["avg"]}%, 95th percentile {ram_stats["p95"]}%',
                    'confidence': 'medium',
                    'priority': 'medium',
                    'estimated_monthly_cost_increase': cost_increase
                })
        
        # Storage optimization (if disk stats available)
        elif (disk_stats and disk_stats['avg'] > 85 and disk_stats['p95'] > 95):
            recommendation.update({
                'action': 'increase_storage',
                'reason': f'High disk utilization: avg {disk_stats["avg"]}%',
                'confidence': 'medium',
                'priority': 'medium'
            })
        
        return recommendation
    
    def _generate_monthly_summary(self, data: pd.DataFrame, recommendations: List[Dict]) -> Dict:
        """Generate overall monthly analysis summary"""
        
        # Count recommendations by action
        action_counts = {}
        total_savings = 0
        total_cost_increase = 0
        
        for rec in recommendations:
            action = rec['action']
            action_counts[action] = action_counts.get(action, 0) + 1
            total_savings += rec.get('estimated_monthly_savings', 0)
            total_cost_increase += rec.get('estimated_monthly_cost_increase', 0)
        
        # Calculate overall utilization
        ec2_data = data[data['Service'] == 'EC2']
        rds_data = data[data['Service'] == 'RDS']
        
        summary = {
            'total_recommendations': len([r for r in recommendations if r['action'] != 'monitor']),
            'action_breakdown': action_counts,
            'estimated_monthly_savings': round(total_savings, 2),
            'estimated_monthly_cost_increase': round(total_cost_increase, 2),
            'net_impact': round(total_savings - total_cost_increase, 2),
            'high_priority_count': len([r for r in recommendations if r['priority'] == 'high']),
            'ec2_instances_analyzed': len(ec2_data['ID'].unique()) if not ec2_data.empty else 0,
            'rds_instances_analyzed': len(rds_data['ID'].unique()) if not rds_data.empty else 0
        }
        
        if not ec2_data.empty:
            summary['avg_ec2_cpu'] = round(pd.to_numeric(ec2_data['CPUUtilization(%)'], errors='coerce').mean(), 2)
            summary['avg_ec2_ram'] = round(pd.to_numeric(ec2_data['RAMUtilization(%)'], errors='coerce').mean(), 2)
        
        if not rds_data.empty:
            summary['avg_rds_cpu'] = round(pd.to_numeric(rds_data['CPUUtilization(%)'], errors='coerce').mean(), 2)
            summary['avg_rds_ram'] = round(pd.to_numeric(rds_data['RAMUtilization(%)'], errors='coerce').mean(), 2)
        
        return summary
    
    def _get_smaller_instance_type(self, current_type: str, service: str) -> str:
        """Get a smaller instance type recommendation"""
        if service == 'EC2':
            downsize_map = {
                # T3a family
                't3a.large': 't3a.medium', 't3a.medium': 't3a.small', 't3a.small': 't3a.micro',
                't3a.xlarge': 't3a.large', 't3a.2xlarge': 't3a.xlarge',
                # T3 family  
                't3.large': 't3.medium', 't3.medium': 't3.small', 't3.small': 't3.micro',
                't3.xlarge': 't3.large', 't3.2xlarge': 't3.xlarge',
                # M5 family
                'm5.large': 't3a.large', 'm5.xlarge': 'm5.large', 'm5.2xlarge': 'm5.xlarge', 'm5.4xlarge': 'm5.2xlarge',
                # C5 family
                'c5.large': 't3a.medium', 'c5.xlarge': 'c5.large', 'c5.2xlarge': 'c5.xlarge',
                # R5 family
                'r5.large': 'm5.large', 'r5.xlarge': 'r5.large', 'r5.2xlarge': 'r5.xlarge'
            }
        else:  # RDS
            downsize_map = {
                # T4g family
                'db.t4g.medium': 'db.t4g.small', 'db.t4g.large': 'db.t4g.medium',
                'db.t4g.xlarge': 'db.t4g.large', 'db.t4g.2xlarge': 'db.t4g.xlarge',
                # T3 family
                'db.t3.medium': 'db.t3.small', 'db.t3.large': 'db.t3.medium',
                'db.t3.xlarge': 'db.t3.large', 'db.t3.2xlarge': 'db.t3.xlarge',
                # M5 family
                'db.m5.large': 'db.t4g.large', 'db.m5.xlarge': 'db.m5.large', 'db.m5.2xlarge': 'db.m5.xlarge',
                # R5 family
                'db.r5.large': 'db.m5.large', 'db.r5.xlarge': 'db.r5.large', 'db.r5.2xlarge': 'db.r5.xlarge'
            }
        
        return downsize_map.get(current_type, current_type)
    
    def _get_larger_instance_type(self, current_type: str, service: str) -> str:
        """Get a larger instance type recommendation"""
        if service == 'EC2':
            upsize_map = {
                # T3a family
                't3a.micro': 't3a.small', 't3a.small': 't3a.medium', 't3a.medium': 't3a.large',
                't3a.large': 't3a.xlarge', 't3a.xlarge': 't3a.2xlarge',
                # T3 family
                't3.micro': 't3.small', 't3.small': 't3.medium', 't3.medium': 't3.large',
                't3.large': 't3.xlarge', 't3.xlarge': 't3.2xlarge',
                # Upgrade from T3 to M5 for consistent performance
                't3a.2xlarge': 'm5.xlarge', 't3.2xlarge': 'm5.xlarge',
                # M5 family
                'm5.large': 'm5.xlarge', 'm5.xlarge': 'm5.2xlarge', 'm5.2xlarge': 'm5.4xlarge',
                # C5 family (for CPU intensive)
                'c5.large': 'c5.xlarge', 'c5.xlarge': 'c5.2xlarge'
            }
        else:  # RDS
            upsize_map = {
                # T4g family
                'db.t4g.micro': 'db.t4g.small', 'db.t4g.small': 'db.t4g.medium', 
                'db.t4g.medium': 'db.t4g.large', 'db.t4g.large': 'db.t4g.xlarge',
                'db.t4g.xlarge': 'db.t4g.2xlarge',
                # T3 family
                'db.t3.micro': 'db.t3.small', 'db.t3.small': 'db.t3.medium',
                'db.t3.medium': 'db.t3.large', 'db.t3.large': 'db.t3.xlarge',
                # Upgrade to M5 for better performance
                'db.t4g.2xlarge': 'db.m5.large', 'db.t3.2xlarge': 'db.m5.large',
                # M5 family
                'db.m5.large': 'db.m5.xlarge', 'db.m5.xlarge': 'db.m5.2xlarge'
            }
        
        return upsize_map.get(current_type, current_type)
    
    def _get_memory_optimized_type(self, current_type: str, service: str) -> str:
        """Get memory optimized instance type recommendation"""
        if service == 'EC2':
            memory_optimized_map = {
                # T3/T3a to R5 (memory optimized)
                't3.medium': 'r5.large', 't3.large': 'r5.large', 't3.xlarge': 'r5.xlarge',
                't3a.medium': 'r5.large', 't3a.large': 'r5.large', 't3a.xlarge': 'r5.xlarge',
                # M5 to R5
                'm5.large': 'r5.large', 'm5.xlarge': 'r5.xlarge', 'm5.2xlarge': 'r5.2xlarge',
                # C5 to R5 (more memory)
                'c5.large': 'r5.large', 'c5.xlarge': 'r5.xlarge'
            }
        else:  # RDS
            memory_optimized_map = {
                # T4g/T3 to R5 (memory optimized)
                'db.t4g.medium': 'db.r5.large', 'db.t4g.large': 'db.r5.large',
                'db.t3.medium': 'db.r5.large', 'db.t3.large': 'db.r5.large',
                # M5 to R5
                'db.m5.large': 'db.r5.large', 'db.m5.xlarge': 'db.r5.xlarge'
            }
        
        return memory_optimized_map.get(current_type, current_type)
    
    def _estimate_cost_savings(self, current_type: str, recommended_type: str, service: str) -> float:
        """Estimate monthly cost savings from downsizing"""
        current_cost = self._get_instance_monthly_cost(current_type, service)
        recommended_cost = self._get_instance_monthly_cost(recommended_type, service)
        return max(0, current_cost - recommended_cost)
    
    def _estimate_cost_increase(self, current_type: str, recommended_type: str, service: str) -> float:
        """Estimate monthly cost increase from upsizing"""
        current_cost = self._get_instance_monthly_cost(current_type, service)
        recommended_cost = self._get_instance_monthly_cost(recommended_type, service)
        return max(0, recommended_cost - current_cost)
    
    def _get_instance_monthly_cost(self, instance_type: str, service: str) -> float:
        """Get approximate monthly cost for instance type (US East pricing)"""
        if service == 'EC2':
            # Approximate EC2 monthly costs (730 hours)
            ec2_costs = {
                't3.nano': 3.80, 't3.micro': 7.66, 't3.small': 15.33, 't3.medium': 30.66, 't3.large': 61.32,
                't3.xlarge': 122.63, 't3.2xlarge': 245.26,
                't3a.nano': 3.28, 't3a.micro': 6.57, 't3a.small': 13.14, 't3a.medium': 26.28, 't3a.large': 52.56,
                't3a.xlarge': 105.12, 't3a.2xlarge': 210.24,
                'm5.large': 70.08, 'm5.xlarge': 140.16, 'm5.2xlarge': 280.32, 'm5.4xlarge': 560.64,
                'c5.large': 62.05, 'c5.xlarge': 124.10, 'c5.2xlarge': 248.20,
                'r5.large': 91.25, 'r5.xlarge': 182.50, 'r5.2xlarge': 365.00, 'r5.4xlarge': 730.00
            }
            return ec2_costs.get(instance_type, 50.0)
        else:  # RDS
            # Approximate RDS monthly costs (730 hours)
            rds_costs = {
                'db.t2.micro': 11.68, 'db.t2.small': 23.36, 'db.t2.medium': 46.72, 'db.t2.large': 93.44,
                'db.t3.micro': 12.41, 'db.t3.small': 24.82, 'db.t3.medium': 49.64, 'db.t3.large': 99.28,
                'db.t3.xlarge': 198.56, 'db.t3.2xlarge': 397.12,
                'db.t4g.micro': 10.22, 'db.t4g.small': 20.44, 'db.t4g.medium': 40.88, 'db.t4g.large': 81.76,
                'db.t4g.xlarge': 163.52, 'db.t4g.2xlarge': 327.04,
                'db.m5.large': 105.12, 'db.m5.xlarge': 210.24, 'db.m5.2xlarge': 420.48,
                'db.r5.large': 136.87, 'db.r5.xlarge': 273.75, 'db.r5.2xlarge': 547.50
            }
            return rds_costs.get(instance_type, 75.0)

    def send_email(self, excel_file: str, df: pd.DataFrame):
        """Send email report"""
        try:
            session = boto3.Session(
                profile_name=self.config['ses_profile'],
                region_name=self.config['ses_region']
            )
            ses_client = session.client('ses')
            
            subject = f"AWS Usage Report - {self.today}"
            
            msg = MIMEMultipart()
            msg['Subject'] = subject
            msg['From'] = f"AWS Infrastructure Report <{self.config['sender_email']}>"
            msg['To'] = self.config['recipient_email']
            
            # Collect stopped instances for email report
            self.logger.info("Collecting stopped/terminated instances for email report...")
            stopped_instances = self.collect_stopped_instances()
            self.logger.info(f"Found {len(stopped_instances)} non-running instances")
            
            html_body = self.create_html_report(df, stopped_instances)
            msg.attach(MIMEText(html_body, 'html'))
            
            with open(excel_file, 'rb') as f:
                excel_data = f.read()
            
            attachment = MIMEApplication(excel_data)
            attachment.add_header('Content-Disposition', 'attachment', filename=excel_file)
            msg.attach(attachment)
            
            response = ses_client.send_raw_email(
                Source=self.config['sender_email'],
                Destinations=[self.config['recipient_email']],
                RawMessage={'Data': msg.as_string()}
            )
            
            self.logger.info(f"Email sent successfully. Message ID: {response['MessageId']}")
            print(f"✓ Email sent to {self.config['recipient_email']}")
            
        except Exception as e:
            self.logger.error(f"Failed to send email: {str(e)}")
            print(f"✗ Failed to send email: {str(e)}")
    
    def create_monthly_excel_report(self, monthly_data: Dict, filename: str = None) -> str:
        """Create Excel report with monthly recommendations"""
        if filename is None:
            filename = f"aws_monthly_recommendations_{datetime.now().strftime('%Y-%m')}.xlsx"
        
        recommendations = monthly_data['recommendations']
        analysis = monthly_data['analysis']
        summary = monthly_data['summary']
        
        # Create recommendations dataframe
        rec_data = []
        for rec in recommendations:
            if rec['action'] != 'monitor':  # Only include actionable recommendations
                rec_data.append({
                    'Instance ID': rec['instance_id'],
                    'Name': rec['name'],
                    'Service': rec['service'],
                    'Current Type': rec['current_type'],
                    'Recommended Type': rec['recommended_type'],
                    'Action': rec['action'],
                    'Priority': rec['priority'],
                    'Confidence': rec['confidence'],
                    'Reason': rec['reason'],
                    'Avg CPU%': rec['cpu_stats']['avg'],
                    'Max CPU%': rec['cpu_stats']['max'],
                    '95th CPU%': rec['cpu_stats']['p95'],
                    'Low CPU Days': f"{rec['cpu_stats']['low_usage_percentage']}%",
                    'Avg RAM%': rec['ram_stats']['avg'],
                    'Max RAM%': rec['ram_stats']['max'],
                    '95th RAM%': rec['ram_stats']['p95'],
                    'Data Points': rec['data_points'],
                    'Monthly Savings ($)': rec.get('estimated_monthly_savings', 0),
                    'Monthly Cost Increase ($)': rec.get('estimated_monthly_cost_increase', 0)
                })
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            if rec_data:
                # Recommendations sheet
                rec_df = pd.DataFrame(rec_data)
                rec_df.to_excel(writer, sheet_name='Recommendations', index=False)
                
                # Summary sheet
                summary_data = [
                    ['Analysis Period (Days)', summary.get('analysis_period_days', 0)],
                    ['Total Instances Analyzed', summary.get('total_instances', 0)],
                    ['EC2 Instances', summary.get('ec2_instances_analyzed', 0)],
                    ['RDS Instances', summary.get('rds_instances_analyzed', 0)],
                    ['', ''],
                    ['Total Recommendations', analysis.get('total_recommendations', 0)],
                    ['High Priority Actions', analysis.get('high_priority_count', 0)],
                    ['', ''],
                    ['Estimated Monthly Savings ($)', analysis.get('estimated_monthly_savings', 0)],
                    ['Estimated Monthly Cost Increase ($)', analysis.get('estimated_monthly_cost_increase', 0)],
                    ['Net Monthly Impact ($)', analysis.get('net_impact', 0)],
                    ['', ''],
                    ['Average EC2 CPU%', analysis.get('avg_ec2_cpu', 'N/A')],
                    ['Average EC2 RAM%', analysis.get('avg_ec2_ram', 'N/A')],
                    ['Average RDS CPU%', analysis.get('avg_rds_cpu', 'N/A')],
                    ['Average RDS RAM%', analysis.get('avg_rds_ram', 'N/A')]
                ]
                
                summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Format the recommendations sheet
                workbook = writer.book
                rec_sheet = writer.sheets['Recommendations']
                
                from openpyxl.styles import PatternFill, Font
                
                # Color code by priority
                high_priority_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                medium_priority_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
                downsize_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                
                for row_idx, row in rec_df.iterrows():
                    row_num = row_idx + 2
                    
                    if row['Priority'] == 'high':
                        for col in range(1, len(rec_df.columns) + 1):
                            rec_sheet.cell(row=row_num, column=col).fill = high_priority_fill
                    elif row['Priority'] == 'medium':
                        for col in range(1, len(rec_df.columns) + 1):
                            rec_sheet.cell(row=row_num, column=col).fill = medium_priority_fill
                    
                    # Highlight savings opportunities
                    if row['Action'] == 'downsize' and row['Monthly Savings ($)'] > 0:
                        savings_col = rec_df.columns.get_loc('Monthly Savings ($)') + 1
                        rec_sheet.cell(row=row_num, column=savings_col).fill = downsize_fill
                        rec_sheet.cell(row=row_num, column=savings_col).font = Font(bold=True)
                
                # Auto-adjust column widths
                for sheet in writer.sheets.values():
                    for column in sheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        sheet.column_dimensions[column_letter].width = adjusted_width
            else:
                # No recommendations sheet
                no_rec_df = pd.DataFrame([['No actionable recommendations found. All instances are optimally sized.']], 
                                       columns=['Message'])
                no_rec_df.to_excel(writer, sheet_name='No Recommendations', index=False)
        
        self.logger.info(f"Monthly Excel report created: {filename}")
        return filename
    
    def create_monthly_html_report(self, monthly_data: Dict) -> str:
        """Create HTML email body for monthly recommendations"""
        recommendations = monthly_data['recommendations']
        analysis = monthly_data['analysis']
        summary = monthly_data['summary']
        
        # Count recommendations by action
        action_counts = analysis.get('action_breakdown', {})
        actionable_recs = [r for r in recommendations if r['action'] != 'monitor']
        high_priority_recs = [r for r in actionable_recs if r['priority'] == 'high']
        
        html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 10px; text-align: center; }}
                .summary {{ background-color: #f8f9fa; padding: 20px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #007bff; }}
                .recommendations {{ margin: 20px 0; }}
                .rec-item {{ background: white; border: 1px solid #dee2e6; border-radius: 8px; padding: 15px; margin: 10px 0; }}
                .high-priority {{ border-left: 4px solid #dc3545; background: #fff5f5; }}
                .medium-priority {{ border-left: 4px solid #ffc107; background: #fffbf0; }}
                .downsize {{ border-left: 4px solid #28a745; background: #f0fff0; }}
                .upsize {{ border-left: 4px solid #fd7e14; background: #fff8f0; }}
                .savings {{ color: #28a745; font-weight: bold; }}
                .cost-increase {{ color: #dc3545; font-weight: bold; }}
                .stats {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; }}
                .stat-card {{ background: white; padding: 15px; border-radius: 8px; text-align: center; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
                .metric {{ font-size: 24px; font-weight: bold; color: #333; }}
                .label {{ color: #666; font-size: 14px; margin-top: 5px; }}
                table {{ width: 100%; border-collapse: collapse; margin: 10px 0; }}
                th, td {{ padding: 8px; text-align: left; border-bottom: 1px solid #ddd; }}
                th {{ background-color: #f8f9fa; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>🎯 AWS Monthly Optimization Report</h1>
                <h2>{datetime.now().strftime('%B %Y')}</h2>
                <p>Comprehensive analysis of your AWS infrastructure usage patterns</p>
            </div>
            
            <div class="summary">
                <h3>📊 Executive Summary</h3>
                <div class="stats">
                    <div class="stat-card">
                        <div class="metric">{summary.get('total_instances', 0)}</div>
                        <div class="label">Instances Analyzed</div>
                    </div>
                    <div class="stat-card">
                        <div class="metric">{analysis.get('total_recommendations', 0)}</div>
                        <div class="label">Recommendations</div>
                    </div>
                    <div class="stat-card">
                        <div class="metric">{analysis.get('high_priority_count', 0)}</div>
                        <div class="label">High Priority</div>
                    </div>
                    <div class="stat-card">
                        <div class="metric">${analysis.get('net_impact', 0):,.2f}</div>
                        <div class="label">Net Monthly Impact</div>
                    </div>
                </div>
                
                <h4>Analysis Period:</h4>
                <ul>
                    <li><strong>{summary.get('analysis_period_days', 0)} days</strong> of usage data</li>
                    <li><strong>{summary.get('data_points', 0)}</strong> total data points collected</li>
                    <li><strong>EC2:</strong> {analysis.get('ec2_instances_analyzed', 0)} instances (avg CPU: {analysis.get('avg_ec2_cpu', 'N/A')}%, avg RAM: {analysis.get('avg_ec2_ram', 'N/A')}%)</li>
                    <li><strong>RDS:</strong> {analysis.get('rds_instances_analyzed', 0)} instances (avg CPU: {analysis.get('avg_rds_cpu', 'N/A')}%, avg RAM: {analysis.get('avg_rds_ram', 'N/A')}%)</li>
                </ul>
            </div>
        """
        
        if actionable_recs:
            # Financial Impact Summary
            total_savings = analysis.get('estimated_monthly_savings', 0)
            total_increases = analysis.get('estimated_monthly_cost_increase', 0)
            
            html += f"""
            <div class="summary">
                <h3>💰 Financial Impact Summary</h3>
                <table>
                    <tr><th>Action Type</th><th>Count</th><th>Monthly Impact</th></tr>
            """
            
            if action_counts.get('downsize', 0) > 0:
                html += f'<tr><td>🔽 Downsize (Cost Savings)</td><td>{action_counts["downsize"]}</td><td class="savings">-${total_savings:,.2f}</td></tr>'
            if action_counts.get('upsize', 0) > 0:
                html += f'<tr><td>🔼 Upsize (Performance)</td><td>{action_counts["upsize"]}</td><td class="cost-increase">+${total_increases:,.2f}</td></tr>'
            if action_counts.get('optimize_memory', 0) > 0:
                html += f'<tr><td>🧠 Memory Optimize</td><td>{action_counts["optimize_memory"]}</td><td class="cost-increase">+${total_increases:,.2f}</td></tr>'
            
            html += f"""
                    <tr style="border-top: 2px solid #333; font-weight: bold;">
                        <td>Net Monthly Impact</td><td>-</td>
                        <td class="{'savings' if analysis.get('net_impact', 0) < 0 else 'cost-increase'}">${analysis.get('net_impact', 0):+,.2f}</td>
                    </tr>
                </table>
            </div>
            """
            
            # High Priority Recommendations
            if high_priority_recs:
                html += '<div class="recommendations"><h3>🚨 High Priority Recommendations</h3>'
                for rec in high_priority_recs[:5]:  # Top 5 high priority
                    priority_class = 'high-priority' if rec['priority'] == 'high' else 'medium-priority'
                    if rec['action'] == 'downsize':
                        priority_class = 'downsize'
                    elif rec['action'] in ['upsize', 'optimize_memory']:
                        priority_class = 'upsize'
                    
                    savings_text = ''
                    if rec.get('estimated_monthly_savings', 0) > 0:
                        savings_text = f' | <span class="savings">Save ${rec["estimated_monthly_savings"]:.2f}/month</span>'
                    elif rec.get('estimated_monthly_cost_increase', 0) > 0:
                        savings_text = f' | <span class="cost-increase">Cost +${rec["estimated_monthly_cost_increase"]:.2f}/month</span>'
                    
                    html += f"""
                    <div class="rec-item {priority_class}">
                        <h4>🎯 {rec['instance_id']} ({rec['name']})</h4>
                        <p><strong>Current:</strong> {rec['current_type']} → <strong>Recommended:</strong> {rec['recommended_type']}</p>
                        <p><strong>Action:</strong> {rec['action'].title()} | <strong>Confidence:</strong> {rec['confidence'].title()}{savings_text}</p>
                        <p><strong>Reason:</strong> {rec['reason']}</p>
                        <p><strong>Usage:</strong> CPU avg {rec['cpu_stats']['avg']}% (95th: {rec['cpu_stats']['p95']}%), RAM avg {rec['ram_stats']['avg']}% (95th: {rec['ram_stats']['p95']}%)</p>
                    </div>
                    """
                html += '</div>'
        else:
            html += """
            <div class="summary">
                <h3>✅ All Systems Optimized!</h3>
                <p>Great news! All your AWS instances appear to be properly sized based on the last 30 days of usage data. No immediate scaling actions are recommended.</p>
                <p>We'll continue monitoring and will alert you if optimization opportunities arise.</p>
            </div>
            """
        
        html += f"""
            <div class="summary">
                <h3>📋 Next Steps</h3>
                <ol>
                    <li><strong>Review High Priority items</strong> - These offer the best ROI</li>
                    <li><strong>Test in staging first</strong> - Validate performance impact</li>
                    <li><strong>Schedule during maintenance windows</strong> - Minimize disruption</li>
                    <li><strong>Monitor post-change</strong> - Verify expected improvements</li>
                </ol>
                <p><em>📧 Detailed recommendations are available in the attached Excel file</em></p>
            </div>
            
            <footer style="text-align: center; margin-top: 30px; padding: 20px; background: #f8f9fa; border-radius: 8px;">
                <p>🤖 Generated automatically by AWS Usage Optimizer on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                <p>Next monthly report: {(datetime.now().replace(day=1) + timedelta(days=32)).replace(day=1).strftime('%Y-%m-%d')}</p>
            </footer>
        </body>
        </html>
        """
        
        return html
    
    def send_monthly_email(self, monthly_data: Dict):
        """Send monthly recommendations email"""
        try:
            session = boto3.Session(
                profile_name=self.config['ses_profile'],
                region_name=self.config['ses_region']
            )
            ses_client = session.client('ses')
            
            # Create Excel report
            excel_file = self.create_monthly_excel_report(monthly_data)
            
            # Create email
            month_year = datetime.now().strftime('%B %Y')
            subject = f"🎯 AWS Monthly Optimization Report - {month_year}"
            
            msg = MIMEMultipart()
            msg['Subject'] = subject
            msg['From'] = f"AWS Optimization Report <{self.config['sender_email']}>"
            msg['To'] = self.config['recipient_email']
            
            # HTML body
            html_body = self.create_monthly_html_report(monthly_data)
            msg.attach(MIMEText(html_body, 'html'))
            
            # Attach Excel file
            with open(excel_file, 'rb') as f:
                excel_data = f.read()
            
            attachment = MIMEApplication(excel_data)
            attachment.add_header('Content-Disposition', 'attachment', filename=excel_file)
            msg.attach(attachment)
            
            # Send email
            response = ses_client.send_raw_email(
                Source=self.config['sender_email'],
                Destinations=[self.config['recipient_email']],
                RawMessage={'Data': msg.as_string()}
            )
            
            self.logger.info(f"Monthly report email sent successfully. Message ID: {response['MessageId']}")
            print(f"✓ Monthly report sent to {self.config['recipient_email']}")
            print(f"✓ Excel file: {excel_file}")
            
        except Exception as e:
            self.logger.error(f"Failed to send monthly report email: {str(e)}")
            print(f"✗ Failed to send monthly report: {str(e)}")

    def run(self):
        """Main execution"""
        self.logger.info(f"Starting AWS Usage Report for {self.today}")
        print(f"\nAWS Usage Reporter")
        print("=" * 50)
        print(f"Date: {self.today}")
        print(f"Profiles: {', '.join(self.profiles)}")
        print(f"Regions: {', '.join(self.config['regions'])}")
        print(f"SSM Timeout: {self.config['ssm_timeout']} seconds")
        print("\n🔧 Collecting data with FIXED SSM commands...")
        
        current_df = self.collect_all_data()
        
        if current_df.empty:
            self.logger.error("No data collected")
            print("✗ No data collected. Check logs for details.")
            return
        
        print(f"✓ Collected data for {len(current_df)} resources")
        
        # Show collection summary
        if 'DataCollection' in current_df.columns:
            ec2_df = current_df[current_df['Service'] == 'EC2']
            ssm_count = len(ec2_df[ec2_df['DataCollection'] == 'SSM+CloudWatch'])
            estimated_count = len(ec2_df[ec2_df['DataCollection'] == 'CloudWatch+Estimated'])
            
            print(f"  - 🎯 Real SSM data: {ssm_count} instances")
            print(f"  - ⚠️  Estimated data: {estimated_count} instances")
            print(f"  - 📊 Success rate: {(ssm_count/(ssm_count+estimated_count)*100) if (ssm_count+estimated_count) > 0 else 0:.1f}%")
        
        historical = self.load_historical_data()
        
        if not (historical.empty if hasattr(historical, 'empty') else len(historical) == 0):
            # Convert to DataFrame if it's a list
            if isinstance(historical, list):
                historical = pd.DataFrame(historical)
                
            yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
            previous_df = historical[historical['Date'] == yesterday]
            current_df = self.calculate_trends(current_df, previous_df)
            print("✓ Calculated trends from previous day")
        else:
            print("ℹ First run - no historical data for trends")
            current_df['CPU_Trend'] = 'N/A'
            current_df['RAM_Trend'] = 'N/A'
            current_df['Disk_Trend'] = 'N/A'
            current_df['CPU_Change'] = 0
            current_df['RAM_Change'] = 0
            current_df['Disk_Change'] = 0
        
        self.save_historical_data(current_df)
        print("✓ Saved historical data")
        
        report_file = self.create_excel_report(current_df)
        print(f"✓ Created Excel report: {report_file}")
        
        self.send_email(report_file, current_df)
        
        # Check if it's the 1st of the month for monthly analysis
        today = datetime.now()
        is_first_of_month = today.day == 1
        
        # Allow manual monthly analysis trigger via environment variable or command line
        force_monthly = os.environ.get('FORCE_MONTHLY_ANALYSIS', 'false').lower() == 'true'
        
        if is_first_of_month or force_monthly:
            print("\n" + "=" * 50)
            print("📊 MONTHLY ANALYSIS TRIGGERED")
            print("=" * 50)
            
            if force_monthly:
                print("🔧 Forced monthly analysis via FORCE_MONTHLY_ANALYSIS=true")
            else:
                print(f"📅 First day of {today.strftime('%B %Y')} - Running monthly analysis")
            
            print("🔍 Analyzing 30-day usage patterns...")
            monthly_data = self.analyze_monthly_trends()
            
            if monthly_data['recommendations']:
                print(f"✓ Found {len([r for r in monthly_data['recommendations'] if r['action'] != 'monitor'])} optimization opportunities")
                print(f"💰 Potential net monthly impact: ${monthly_data['analysis'].get('net_impact', 0):+,.2f}")
                
                high_priority = len([r for r in monthly_data['recommendations'] if r['priority'] == 'high'])
                if high_priority > 0:
                    print(f"🚨 {high_priority} high-priority recommendations require immediate attention!")
                
                print("📧 Sending monthly optimization report...")
                self.send_monthly_email(monthly_data)
            else:
                print("✅ No optimization opportunities found - all instances are properly sized!")
                
                # Send "all good" monthly report
                monthly_data['analysis']['message'] = "All instances optimally sized"
                print("📧 Sending monthly status report...")
                self.send_monthly_email(monthly_data)

        print("\n" + "=" * 50)
        print("🎯 AWS Usage Report complete!")


if __name__ == "__main__":
    import sys
    
    # Check for command line arguments
    if len(sys.argv) > 1 and sys.argv[1] == '--monthly':
        os.environ['FORCE_MONTHLY_ANALYSIS'] = 'true'
        print("🔧 Monthly analysis forced via --monthly flag")
    
    reporter = AWSFixedReporter()
    reporter.run()